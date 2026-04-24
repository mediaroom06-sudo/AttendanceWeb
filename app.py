from flask import Flask, render_template, request, redirect, send_file
import face_recognition
import os
import sqlite3
import numpy as np
from datetime import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

app = Flask(__name__)

UPLOAD_FOLDER = '/tmp/uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ---------------- DATABASE ----------------
conn = sqlite3.connect('database.db', check_same_thread=False)

def query(sql, params=(), fetch=False):
    cur = conn.cursor()
    cur.execute(sql, params)
    data = cur.fetchall() if fetch else None
    conn.commit()
    cur.close()
    return data

# ---------------- TABLES ----------------
query("""
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    encoding BLOB,
    group_name TEXT
)
""")

query("""
CREATE TABLE IF NOT EXISTS attendance (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    date TEXT,
    status TEXT
)
""")

# ---------------- HOME ----------------
@app.route('/')
def index():
    return render_template('index.html')

# ---------------- REGISTER ----------------
@app.route('/register', methods=['GET', 'POST'])
def register():

    if request.method == 'POST':

        name = request.form['name']
        group = request.form['group']
        file = request.files['image']

        if file.filename == "":
            return "No file selected"

        path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(path)

        image = face_recognition.load_image_file(path)
        encodings = face_recognition.face_encodings(image)

        if not encodings:
            return "No face detected"

        encoding = encodings[0]

        users = query("SELECT name, encoding FROM users", fetch=True)
        known_encodings = [np.frombuffer(u[1], dtype=np.float64) for u in users]

        if len(known_encodings) > 0:
            distances = face_recognition.face_distance(known_encodings, encoding)
            best = np.argmin(distances)

            if distances[best] < 0.40:
                return "Face already registered"

        query(
            "INSERT INTO users (name, encoding, group_name) VALUES (?, ?, ?)",
            (name, encoding.tobytes(), group)
        )

        return redirect('/register')

    return render_template('register.html')

# ---------------- UPLOAD ----------------
@app.route('/upload', methods=['POST', 'GET'])
def upload():

    if request.method == 'POST':

        file = request.files.get('image')
        upload_time = request.form.get('time', "08:00")

        if not file or file.filename == "":
            return "No file uploaded"

        path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(path)

        img = face_recognition.load_image_file(path)
        faces = face_recognition.face_encodings(img)

        users = query("SELECT name, encoding FROM users", fetch=True)

        known_names = [u[0] for u in users]
        known_encodings = [np.frombuffer(u[1], dtype=np.float64) for u in users]

        already_marked = set()

        # ✅ NEW STATUS LOGIC (FIXED)
        def get_status():
            try:
                t = datetime.strptime(upload_time, "%H:%M").time()
            except:
                return "RED"

            # Morning
            if datetime.strptime("08:00", "%H:%M").time() <= t <= datetime.strptime("09:00", "%H:%M").time():
                return "GREEN"

            if datetime.strptime("09:01", "%H:%M").time() <= t <= datetime.strptime("10:00", "%H:%M").time():
                return "ORANGE"

            # Evening
            if datetime.strptime("18:00", "%H:%M").time() <= t <= datetime.strptime("19:30", "%H:%M").time():
                return "GREEN"

            if datetime.strptime("19:31", "%H:%M").time() <= t <= datetime.strptime("20:30", "%H:%M").time():
                return "ORANGE"

            return "RED"

        # ✅ Create datetime from user input
        today = datetime.now().strftime("%Y-%m-%d")
        full_datetime = f"{today} {upload_time}:00"

        for face in faces:

            if len(known_encodings) == 0:
                continue

            distances = face_recognition.face_distance(known_encodings, face)
            best = np.argmin(distances)

            if distances[best] < 0.5:

                name = known_names[best]

                if name in already_marked:
                    continue

                already_marked.add(name)

                # ✅ Check if already marked today
                exists = query(
                    "SELECT * FROM attendance WHERE name=? AND date LIKE ?",
                    (name, f"{today}%"),
                    fetch=True
                )

                if not exists:
                    query(
                        "INSERT INTO attendance (name, date, status) VALUES (?, ?, ?)",
                        (name, full_datetime, get_status())
                    )

        return redirect('/attendance')

    return render_template('upload.html')
# ---------------- ATTENDANCE ----------------
@app.route('/attendance')
def attendance():

    users = query("SELECT name FROM users", fetch=True)
    all_names = sorted(set([u[0] for u in users]))

    records = query("SELECT name, date, status FROM attendance", fetch=True)

    dates = sorted(set([
        datetime.strptime(r[1][:10], "%Y-%m-%d").strftime("%d/%m/%y")
        for r in records
    ]))

    attendance_map = {}

    for name in all_names:
        attendance_map[name] = {}
        for d in dates:
            attendance_map[name][d] = {
                "status": "RED",
                "time": "-"
            }

    for name, dt, status in records:
        formatted = datetime.strptime(dt[:10], "%Y-%m-%d").strftime("%d/%m/%y")

        # ✅ Safe time extraction
        try:
            time_str = datetime.strptime(dt, "%Y-%m-%d %H:%M:%S").strftime("%H:%M")
        except:
            time_str = dt.split(" ")[1][:5] if " " in dt else "-"

        attendance_map[name][formatted] = {
            "status": status,
            "time": time_str
        }

    users_group = query("SELECT name, group_name FROM users", fetch=True)

    groups = {}
    for name, group in users_group:
        groups.setdefault(group, []).append(name)

    return render_template(
        "attendance.html",
        names=all_names,
        dates=dates,
        attendance=attendance_map,
        groups=groups
    )

# ---------------- DOWNLOAD EXCEL ----------------
@app.route('/download-excel')
def download_excel():

    users = query("SELECT name, group_name FROM users", fetch=True)
    records = query("SELECT name, date, status FROM attendance", fetch=True)

    name_to_group = {u[0]: u[1] for u in users}
    all_names = list(name_to_group.keys())

    dates = sorted(set([
        datetime.strptime(r[1][:10], "%Y-%m-%d").strftime("%d/%m/%y")
        for r in records
    ]))

    wb = Workbook()
    wb.remove(wb.active)

    green = PatternFill("solid", fgColor="2ecc71")
    orange = PatternFill("solid", fgColor="f39c12")
    red = PatternFill("solid", fgColor="e74c3c")

    for group in set(name_to_group.values()):

        ws = wb.create_sheet(title=str(group)[:30])

        ws.cell(row=1, column=1, value="Name").font = Font(bold=True)

        for col, d in enumerate(dates, start=2):
            ws.cell(row=1, column=col, value=d).font = Font(bold=True)

        row_num = 2

        for name in all_names:

            if name_to_group.get(name) != group:
                continue

            ws.cell(row=row_num, column=1, value=name)

            for col, d in enumerate(dates, start=2):

                status = "RED"

                for r in records:
                    formatted = datetime.strptime(r[1][:10], "%Y-%m-%d").strftime("%d/%m/%y")
                    if r[0] == name and formatted == d:
                        status = r[2]

                cell = ws.cell(row=row_num, column=col)

                if status == "GREEN":
                    cell.fill = green
                elif status == "ORANGE":
                    cell.fill = orange
                else:
                    cell.fill = red

            row_num += 1

        ws.freeze_panes = "B2"

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="attendance.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---------------- DELETE USER ----------------
@app.route('/delete-user/<path:name>')
def delete_user(name):
    query("DELETE FROM users WHERE name=?", (name,))
    query("DELETE FROM attendance WHERE name=?", (name,))
    return redirect('/attendance')

# ---------------- EDIT USER ----------------
@app.route('/edit-user/<path:name>', methods=['GET', 'POST'])
def edit_user(name):

    if request.method == 'POST':
        new_name = request.form['name']
        new_group = request.form['group']

        # update user table
        query(
            "UPDATE users SET name=?, group_name=? WHERE name=?",
            (new_name, new_group, name)
        )

        # update attendance records too
        query(
            "UPDATE attendance SET name=? WHERE name=?",
            (new_name, name)
        )

        return redirect('/attendance')

    # get current user info
    user = query(
        "SELECT name, group_name FROM users WHERE name=?",
        (name,),
        fetch=True
    )

    if not user:
        return "User not found"

    return render_template('edit_user.html', name=user[0][0], group=user[0][1])
# ---------------- RUN ----------------
if __name__ == '__main__':
    app.run()